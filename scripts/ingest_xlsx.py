#!/usr/bin/env python3
"""
ARCHIVED — historical one-shot bootstrap.

This script seeded the YAML catalogue from the original
business-capability-catalogue.xlsx spreadsheet. The seed has been
committed; the spreadsheet is no longer in the repository and is no
longer a source of truth. Re-running this script is not part of the
governance workflow.

Per governance.md, all new capabilities and metadata changes are
introduced as YAML edits under catalogue/, gated by `npm run lint`
and reviewed via pull request. Do not use this script to import new
data.
"""
from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed; run `pip install openpyxl`")

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "business-capability-catalogue.xlsx"
CATALOGUE = REPO / "catalogue"


def slugify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return s


def quote(s: str) -> str:
    """Single-line YAML string. Wrap in double quotes if it contains chars
    that would otherwise need escaping or are ambiguous."""
    if s == "":
        return '""'
    if re.search(r'[:#&*!|>%@`,\[\]\{\}\'"\n]', s) or s[0] in " -?":
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return s


def block_paragraph(text: str, indent: int) -> str:
    """Render `text` as a YAML block scalar (`|`) with the given indent."""
    pad = " " * indent
    lines = text.strip().split("\n")
    return "|\n" + "\n".join(pad + ln for ln in lines)


def emit_l1_yaml(l1, l2s) -> str:
    out = []
    out.append("# " + l1["name"])
    out.append(f"id: {l1['id']}")
    out.append(f"name: {quote(l1['name'])}")
    out.append("level: 1")
    out.append(f"industry: {quote(l1['industry'])}")
    out.append(f"description: {block_paragraph(l1['description'], indent=2)}")
    if not l2s:
        out.append("children: []")
    else:
        out.append("children:")
        for child in l2s:
            out.append(f"  - id: {child['id']}")
            out.append(f"    name: {quote(child['name'])}")
            out.append("    level: 2")
            out.append(f"    industry: {quote(child['industry'])}")
            out.append(
                f"    description: {block_paragraph(child['description'], indent=6)}"
            )
            out.append("    children: []")
    return "\n".join(out) + "\n"


def cap_id_sort_key(cid: str) -> list[int]:
    return [int(x) for x in cid.replace("BC-", "").split(".")]


def main() -> None:
    if not XLSX.exists():
        sys.exit(f"Missing source spreadsheet: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    cap_ws = wb["Capabilities"]
    vs_ws = wb["Value Streams"]

    cap_rows = list(cap_ws.iter_rows(values_only=True))
    cap_header = cap_rows[0]
    idx = {name: i for i, name in enumerate(cap_header)}

    capabilities = []
    for row in cap_rows[1:]:
        cap = {
            "id": row[idx["ID"]],
            "name": row[idx["Name"]],
            "level_str": row[idx["Level"]],
            "parent_id": row[idx["Parent ID"]],
            "industry": (row[idx["Industry"]] or "").strip(),
            "description": (row[idx["Definition"]] or "").strip(),
        }
        if not cap["id"] or not cap["name"]:
            continue
        capabilities.append(cap)

    by_parent: dict[str | None, list[dict]] = defaultdict(list)
    for c in capabilities:
        by_parent[c["parent_id"]].append(c)
    for k in by_parent:
        by_parent[k].sort(key=lambda c: cap_id_sort_key(c["id"]))

    l1s = [c for c in capabilities if c["level_str"] == "L1"]
    l1s.sort(key=lambda c: cap_id_sort_key(c["id"]))

    slug_seen: dict[str, str] = {}
    written: list[str] = []

    # Wipe existing L1-*.yaml files (the demo BC-2 / BC-3 set).
    for old in CATALOGUE.glob("L1-*.yaml"):
        old.unlink()

    for l1 in l1s:
        slug = slugify(l1["name"])
        if slug in slug_seen:
            sys.exit(
                f"L1 slug collision: '{slug}' from {l1['id']} also from {slug_seen[slug]}"
            )
        slug_seen[slug] = l1["id"]
        l2s = by_parent.get(l1["id"], [])
        path = CATALOGUE / f"L1-{slug}.yaml"
        path.write_text(emit_l1_yaml(l1, l2s), encoding="utf-8")
        written.append(path.name)

    # _index.yaml
    index_path = CATALOGUE / "_index.yaml"
    lines = [
        "# Registry of L1 catalogue files. Lint enforces that every file in catalogue/",
        "# (other than this index and underscore-prefixed meta files) is listed here,",
        "# and every entry resolves to an existing file.",
        "files:",
    ]
    for name in written:
        lines.append(f"  - {name}")
    index_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # _value-streams.yaml
    vs_rows = list(vs_ws.iter_rows(values_only=True))
    vs_header = vs_rows[0]
    vidx = {name: i for i, name in enumerate(vs_header)}

    by_stream: dict[str, list[dict]] = defaultdict(list)
    for row in vs_rows[1:]:
        if not row[vidx["Value Stream"]]:
            continue
        by_stream[row[vidx["Value Stream"]]].append(
            {
                "stage_order": row[vidx["Stage Order"]],
                "stage_name": (row[vidx["Stage Name"]] or "").strip(),
                "capability_id": (row[vidx["Capability ID"]] or "").strip(),
                "industry_variant": (row[vidx["Industry Variant"]] or "").strip(),
                "notes": (row[vidx["Notes"]] or "").strip(),
            }
        )

    out = [
        "# Value Streams (orthogonal artefact, not part of the capability hierarchy).",
        "# Each entry links a capability to one stage of an end-to-end flow.",
        "value_streams:",
    ]
    for stream in sorted(by_stream):
        out.append(f"  - name: {quote(stream)}")
        out.append("    stages:")
        for entry in sorted(
            by_stream[stream],
            key=lambda e: (e["stage_order"] or 0, e["capability_id"]),
        ):
            out.append(f"      - stage_order: {entry['stage_order']}")
            out.append(f"        stage_name: {quote(entry['stage_name'])}")
            out.append(f"        capability_id: {entry['capability_id']}")
            if entry["industry_variant"] and entry["industry_variant"] != "All":
                out.append(
                    f"        industry_variant: {quote(entry['industry_variant'])}"
                )
            if entry["notes"]:
                out.append(f"        notes: {quote(entry['notes'])}")

    (CATALOGUE / "_value-streams.yaml").write_text(
        "\n".join(out) + "\n", encoding="utf-8"
    )

    print(
        f"Wrote {len(written)} L1 file(s), {sum(len(v) for v in by_stream.values())} "
        f"value-stream link(s) across {len(by_stream)} stream(s) → catalogue/"
    )


if __name__ == "__main__":
    main()
